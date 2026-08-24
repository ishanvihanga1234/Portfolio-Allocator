"""
Private Wealth Portfolio Allocation System
============================================
Sri Lanka equity + G-sec allocation engine.

Daily-updated inputs (read from disk, no hardcoding):
  - Portfolio.xlsx          -> today's stock prices, 12-month price forecasts, T-bill rates
  - Full_Dataset_5_Stocks.xlsx -> historical monthly prices/returns for covariance (risk) estimation

Per-investor runtime inputs:
  - tenor_month   : the calendar month (Jan..Dec, 2025) that matches the investor's real horizon
                     (i.e. the G-sec whose maturity/forecast month lines up with when they exit)
  - investment_amount : LKR amount the investor wants to deploy
  - preference_list   : optional list of tickers the investor prefers (subset of the 5-stock basket)

Pipeline (mirrors the advisor's framework):
  1. Screen: for the chosen tenor month, keep only stocks whose forecast horizon return beats
     the matching T-bill rate. If a preference list is given, use it ONLY if at least one
     preferred stock survives the screen; otherwise ignore the preference and use the full
     screened basket.
  2. Expected returns (mu) for surviving stocks = forecast horizon return (from Table 2), NOT
     historical average -> "forecasting model output, no screening" per the framework.
  3. Covariance (Sigma) estimated from historical monthly returns (Full_Dataset), scaled to the
     investor's horizon length in months (assumes iid monthly returns).
  4. Markowitz tangency (max-Sharpe) portfolio from (mu, Sigma, Rf).
  5. Risk aversion (gamma) is NOT asked from the investor as a hypothetical gamble. It is backed
     out automatically:
        - Guarantee amount = what the investment would be worth if fully in the risk-free G-sec
          for the horizon: FV_safe = W0 * (1 + Rf_tenor)
        - Risky 50/50 gamble outcomes = tangency portfolio +-1 std dev around its expected
          horizon return:
              Upside   = W0 * (1 + mu_p + sigma_p)
              Downside = W0 * (1 + mu_p - sigma_p)
        - Solve  0.5*U(Upside) + 0.5*U(Downside) = U(FV_safe)  for gamma, CRRA utility,
          via bisection (same method as the advisor's manual CE elicitation, just automated
          with market-implied numbers instead of a hypothetical question).
  6. Two-fund separation (Merton):  y* = (mu_p - Rf_tenor) / (gamma * sigma_p^2)
        y* = weight in the risky (tangency) portfolio, (1 - y*) = weight in G-sec.
        Per-stock equity weight = y* * tangency_weight_i.
  7. Report: portfolio expected return, portfolio volatility, G-sec weight, per-stock weights.
"""

import numpy as np
import pandas as pd
import openpyxl
from dataclasses import dataclass, field


MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
TICKERS = ["JKH", "COMB", "SAMP", "CTC", "HAYL"]


# ---------------------------------------------------------------------------
# 1. Load daily market data (Portfolio.xlsx)
# ---------------------------------------------------------------------------

def load_market_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    today_price = {}
    for r in range(3, 8):
        name = ws.cell(row=r, column=2).value
        price = ws.cell(row=r, column=3).value
        today_price[name] = price

    # forecast prices: rows 3-14 = Jan..Dec, cols F,G,H,I,J = JKH,COMB,SAMP,CTC,HAYL
    fc_cols = {"JKH": 6, "COMB": 7, "SAMP": 8, "CTC": 9, "HAYL": 10}
    forecast = {t: {} for t in TICKERS}
    for i, month in enumerate(MONTHS):
        r = 3 + i
        for t, c in fc_cols.items():
            forecast[t][month] = ws.cell(row=r, column=c).value

    # T-bill: rows 3-14 = Jan..Dec, col O = rate (%), col N = maturity days
    tbill_rate = {}
    tbill_maturity_days = {}
    for i, month in enumerate(MONTHS):
        r = 3 + i
        tbill_rate[month] = ws.cell(row=r, column=15).value       # O
        tbill_maturity_days[month] = ws.cell(row=r, column=14).value  # N

    return {
        "today_price": today_price,
        "forecast": forecast,
        "tbill_rate": tbill_rate,
        "tbill_maturity_days": tbill_maturity_days,
    }


# ---------------------------------------------------------------------------
# 2. Load historical monthly data (Full_Dataset_5_Stocks.xlsx) -> covariance
# ---------------------------------------------------------------------------

def load_historical_returns(path):
    """Return a DataFrame (index=Month, columns=tickers) of monthly % returns,
    restricted to months common to all 5 stocks (handles CTC's trading gaps)."""
    xls = pd.ExcelFile(path)
    series = {}
    for t in TICKERS:
        df = pd.read_excel(xls, sheet_name=t, usecols=["Month", "Return_%"])
        df = df.dropna(subset=["Return_%"])
        df = df.set_index("Month")["Return_%"]
        series[t] = df
    ret_df = pd.DataFrame(series).dropna(how="any")
    return ret_df  # monthly % returns, e.g. 1.87 means 1.87%


# ---------------------------------------------------------------------------
# 3. Screening logic (replicates the "Final Table" formula, computed fresh)
# ---------------------------------------------------------------------------

def screen_stocks(market, tenor_month):
    """For the chosen tenor month, return dict{ticker: horizon_return_%} for
    stocks whose forecast horizon return beats the matching T-bill rate."""
    rf = market["tbill_rate"][tenor_month]
    eligible = {}
    all_returns = {}
    for t in TICKERS:
        p0 = market["today_price"][t]
        pf = market["forecast"][t][tenor_month]
        horizon_return_pct = (pf - p0) / p0 * 100.0
        all_returns[t] = horizon_return_pct
        if horizon_return_pct > rf:
            eligible[t] = horizon_return_pct
    return eligible, all_returns, rf


def apply_preference(eligible, preference_list):
    """Use the preference list only if at least one preferred ticker survives
    the screen; otherwise fall back to the full eligible screened set."""
    if not preference_list:
        return eligible, "no preference given -> using full screened basket"

    preferred_eligible = {t: r for t, r in eligible.items() if t in preference_list}
    if preferred_eligible:
        return preferred_eligible, f"preference list honoured -> {list(preferred_eligible)}"
    else:
        return eligible, "none of the preferred stocks beat the G-sec rate for this tenor -> preference ignored, using full screened basket"


# ---------------------------------------------------------------------------
# 4. Markowitz tangency (max-Sharpe) portfolio
# ---------------------------------------------------------------------------

def tangency_portfolio(mu_pct, cov_pct, rf_pct):
    """
    mu_pct  : np.array of expected horizon returns (%), for surviving stocks
    cov_pct : covariance matrix of horizon returns (%^2), same order
    rf_pct  : risk-free horizon return (%)
    Returns (weights, mu_p_pct, sigma_p_pct)
    Unconstrained (allows short positions) tangency formula: w ~ Sigma^-1 (mu - rf)
    """
    excess = mu_pct - rf_pct
    inv_cov = np.linalg.inv(cov_pct)
    raw_w = inv_cov @ excess
    w = raw_w / raw_w.sum()
    mu_p = w @ mu_pct
    sigma_p = np.sqrt(w @ cov_pct @ w)
    return w, mu_p, sigma_p


# ---------------------------------------------------------------------------
# 5. CRRA utility + bisection solver for gamma
# ---------------------------------------------------------------------------

def crra_utility(w, gamma):
    if gamma == 1.0:
        return np.log(w)
    return (w ** (1 - gamma)) / (1 - gamma)


def solve_gamma(guarantee, upside, downside, lo=0.05, hi=30.0, tol=1e-7, max_iter=200):
    """Bisection solve for gamma s.t. 0.5U(upside)+0.5U(downside) = U(guarantee).

    NOTE: for large gamma, W**(1-gamma) underflows toward 0 in floating point, so
    f(gamma) can look like ~0 far from the true root. We must NOT stop on
    abs(f_mid) < tol for that reason -- only the interval width (hi - lo) is a
    reliable convergence criterion here, since the sign of f stays correct even
    when its magnitude underflows.
    """
    def f(gamma):
        return 0.5 * crra_utility(upside, gamma) + 0.5 * crra_utility(downside, gamma) - crra_utility(guarantee, gamma)

    f_lo, f_hi = f(lo), f(hi)
    if f_lo == 0:
        return lo
    if f_hi == 0:
        return hi
    if f_lo * f_hi > 0:
        # widen the search range a bit before giving up
        for hi_try in (50.0, 80.0, 120.0):
            f_hi = f(hi_try)
            if f_lo * f_hi <= 0:
                hi = hi_try
                break
        else:
            raise ValueError("Could not bracket a root for gamma in [0.05, 120]. "
                              "Check the guarantee/upside/downside inputs.")

    for _ in range(max_iter):
        if (hi - lo) < tol:
            break
        mid = (lo + hi) / 2
        f_mid = f(mid)
        if f_lo * f_mid <= 0:
            hi = mid
            f_hi = f_mid
        else:
            lo = mid
            f_lo = f_mid
    return (lo + hi) / 2


# ---------------------------------------------------------------------------
# 6. Full pipeline
# ---------------------------------------------------------------------------

@dataclass
class AllocationResult:
    tenor_month: str
    rf_pct: float
    screened: dict
    preference_note: str
    used_stocks: list
    tangency_weights: dict
    mu_p_pct: float
    sigma_p_pct: float
    guarantee_amount: float
    upside_amount: float
    downside_amount: float
    gamma: float
    y_star: float
    gsec_weight: float
    equity_weights: dict
    portfolio_expected_return_pct: float
    portfolio_volatility_pct: float
    portfolio_expected_value: float


def run_allocation(portfolio_xlsx, dataset_xlsx, tenor_month, investment_amount,
                    preference_list=None):
    if tenor_month not in MONTHS:
        raise ValueError(f"tenor_month must be one of {MONTHS}")

    market = load_market_data(portfolio_xlsx)
    hist = load_historical_returns(dataset_xlsx)

    # tenor length in months (Jan=1 ... i.e. months elapsed from the Jan 1 meeting date)
    tenor_months = MONTHS.index(tenor_month) + 1

    # --- Step 1: screen + preference -----------------------------------
    eligible, all_returns, rf_pct = screen_stocks(market, tenor_month)
    if not eligible:
        # nobody beats the bond -> 100% G-sec
        guarantee = investment_amount * (1 + rf_pct / 100.0)
        return AllocationResult(
            tenor_month=tenor_month, rf_pct=rf_pct, screened={},
            preference_note="no stock beat the G-sec rate for this tenor -> 100% G-sec",
            used_stocks=[], tangency_weights={}, mu_p_pct=0.0, sigma_p_pct=0.0,
            guarantee_amount=guarantee, upside_amount=guarantee, downside_amount=guarantee,
            gamma=float("nan"), y_star=0.0, gsec_weight=1.0, equity_weights={},
            portfolio_expected_return_pct=rf_pct, portfolio_volatility_pct=0.0,
            portfolio_expected_value=guarantee,
        )

    used, pref_note = apply_preference(eligible, preference_list)
    used_stocks = list(used.keys())
    mu_pct = np.array([used[t] for t in used_stocks])

    # --- Step 2: covariance from historical monthly data, scaled to horizon
    sub_hist = hist[used_stocks]
    monthly_cov = sub_hist.cov().values  # in %^2 (since Return_% column already in percent units)
    cov_pct = monthly_cov * tenor_months  # scale variance/covariance by horizon length (iid months)

    # --- Step 3: Markowitz tangency portfolio ---------------------------
    w, mu_p, sigma_p = tangency_portfolio(mu_pct, cov_pct, rf_pct)
    tangency_weights = dict(zip(used_stocks, w))

    # --- Step 4: guarantee / upside / downside dollar amounts -----------
    guarantee = investment_amount * (1 + rf_pct / 100.0)
    upside = investment_amount * (1 + (mu_p + sigma_p) / 100.0)
    downside = investment_amount * (1 + (mu_p - sigma_p) / 100.0)
    downside = max(downside, 1e-6)  # guard against non-positive wealth for CRRA utility

    # --- Step 5: solve gamma ---------------------------------------------
    gamma = solve_gamma(guarantee, upside, downside)

    # --- Step 6: two-fund separation (Merton) -----------------------------
    # mu_p, rf_pct, sigma_p are all in percentage-point units (e.g. 14.84 = 14.84%).
    # The Merton formula y* = (mu_p - Rf) / (gamma * sigma_p^2) needs *decimal*
    # units (0.1484, not 14.84) for excess return and variance to cancel out
    # correctly -- so divide the percentage-point values by 100 before using them.
    excess_decimal = (mu_p - rf_pct) / 100.0
    var_decimal = (sigma_p / 100.0) ** 2
    y_star = excess_decimal / (gamma * var_decimal)
    y_star_clamped = min(max(y_star, 0.0), 1.0)  # long-only, no leverage
    gsec_weight = 1 - y_star_clamped
    equity_weights = {t: y_star_clamped * tangency_weights[t] for t in used_stocks}

    port_ret = y_star_clamped * mu_p + gsec_weight * rf_pct
    port_vol = y_star_clamped * sigma_p
    port_value = investment_amount * (1 + port_ret / 100.0)

    return AllocationResult(
        tenor_month=tenor_month, rf_pct=rf_pct, screened=eligible,
        preference_note=pref_note, used_stocks=used_stocks,
        tangency_weights=tangency_weights, mu_p_pct=mu_p, sigma_p_pct=sigma_p,
        guarantee_amount=guarantee, upside_amount=upside, downside_amount=downside,
        gamma=gamma, y_star=y_star, gsec_weight=gsec_weight,
        equity_weights=equity_weights,
        portfolio_expected_return_pct=port_ret, portfolio_volatility_pct=port_vol,
        portfolio_expected_value=port_value,
    )


# ---------------------------------------------------------------------------
# 7. Pretty print
# ---------------------------------------------------------------------------

def print_report(res: AllocationResult, investment_amount: float):
    print("=" * 60)
    print(f"  PORTFOLIO ALLOCATION REPORT — tenor: {res.tenor_month}")
    print("=" * 60)
    print(f"Investment amount        : Rs. {investment_amount:,.2f}")
    print(f"G-sec (T-bill) rate      : {res.rf_pct:.2f}%")
    print(f"Preference handling      : {res.preference_note}")
    print(f"Stocks used in portfolio : {', '.join(res.used_stocks) if res.used_stocks else 'None'}")
    if res.used_stocks:
        print(f"Tangency portfolio return: {res.mu_p_pct:.2f}%")
        print(f"Tangency portfolio vol   : {res.sigma_p_pct:.2f}%")
        print(f"Guarantee (G-sec) amount : Rs. {res.guarantee_amount:,.2f}")
        print(f"Risky upside amount      : Rs. {res.upside_amount:,.2f}")
        print(f"Risky downside amount    : Rs. {res.downside_amount:,.2f}")
        print(f"Implied gamma (CRRA)     : {res.gamma:.4f}")
        print(f"y* (equity weight, raw)  : {res.y_star:.4f}")
    print("-" * 60)
    print(f"FINAL G-sec weight       : {res.gsec_weight*100:.2f}%")
    for t, w in res.equity_weights.items():
        print(f"FINAL {t} weight{' '*(10-len(t))}: {w*100:.2f}%")
    print("-" * 60)
    print(f"Portfolio expected return: {res.portfolio_expected_return_pct:.2f}%")
    print(f"Portfolio volatility     : {res.portfolio_volatility_pct:.2f}%")
    print(f"Portfolio expected value : Rs. {res.portfolio_expected_value:,.2f}")
    print("=" * 60)
