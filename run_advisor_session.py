"""
Advisor-facing session runner.

Usage (interactive):
    python run_advisor_session.py

Usage (non-interactive / scripted):
    python run_advisor_session.py --tenor June --amount 1000000 --preference JKH,HAYL
    python run_advisor_session.py --tenor December --amount 2500000

Reads the two daily-updated workbooks from the same folder (or paths given with
--portfolio / --dataset), asks (or takes as args) the investor's tenor,
investment amount and optional preference list, runs the allocation engine,
prints the report to screen, and writes a formatted Excel report for the
investor's file.
"""

import argparse
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from portfolio_engine import run_allocation, print_report, MONTHS, TICKERS


def parse_args():
    p = argparse.ArgumentParser(description="Run one investor allocation session.")
    p.add_argument("--portfolio", default="Portfolio.xlsx",
                    help="Path to the daily-updated Portfolio.xlsx")
    p.add_argument("--dataset", default="Full_Dataset_5_Stocks.xlsx",
                    help="Path to Full_Dataset_5_Stocks.xlsx")
    p.add_argument("--tenor", default=None, help="Tenor month, e.g. June")
    p.add_argument("--amount", type=float, default=None, help="Investment amount (LKR)")
    p.add_argument("--preference", default=None,
                    help="Comma-separated preferred tickers, e.g. JKH,HAYL (optional)")
    p.add_argument("--investor-name", default="Investor", help="Name for the report title")
    p.add_argument("--outdir", default="/mnt/user-data/outputs",
                    help="Folder to write the Excel report into")
    return p.parse_args()


def ask_interactive():
    print("Tenor options:", ", ".join(MONTHS))
    tenor = input("Investor's expected tenor (month name, e.g. June): ").strip().title()
    while tenor not in MONTHS:
        tenor = input(f"Please type one of {MONTHS}: ").strip().title()

    while True:
        try:
            amount = float(input("Investment amount (LKR): ").strip().replace(",", ""))
            break
        except ValueError:
            print("Please enter a number.")

    pref_raw = input(f"Preferred stocks (comma-separated from {TICKERS}, or press Enter for none): ").strip()
    preference = [t.strip().upper() for t in pref_raw.split(",") if t.strip()] if pref_raw else None

    name = input("Investor name (for the report, optional): ").strip() or "Investor"
    return tenor, amount, preference, name


def build_excel_report(res, investment_amount, investor_name, tenor, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Report"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    section_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    pct_fmt = "0.00%"
    money_fmt = "#,##0.00"

    r = 1
    ws.cell(row=r, column=1, value="Private Portfolio Allocation Report").font = title_font
    r += 1
    ws.cell(row=r, column=1, value=f"Investor: {investor_name}").font = bold
    r += 1
    ws.cell(row=r, column=1, value=f"Date generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    r += 1
    ws.cell(row=r, column=1, value=f"Investment horizon (tenor): {tenor}")
    r += 1
    ws.cell(row=r, column=1, value="Investment amount (LKR):")
    ws.cell(row=r, column=2, value=investment_amount).number_format = money_fmt
    r += 2

    ws.cell(row=r, column=1, value="Model Inputs & Diagnostics").font = bold
    ws.cell(row=r, column=1).fill = section_fill
    r += 1
    rows = [
        ("G-sec (T-bill) rate for this tenor", res.rf_pct / 100, pct_fmt),
        ("Preference handling", res.preference_note, None),
        ("Stocks used", ", ".join(res.used_stocks) if res.used_stocks else "None (100% G-sec)", None),
        ("Tangency portfolio expected return", res.mu_p_pct / 100 if res.used_stocks else None, pct_fmt),
        ("Tangency portfolio volatility", res.sigma_p_pct / 100 if res.used_stocks else None, pct_fmt),
        ("Guarantee amount (100% G-sec)", res.guarantee_amount, money_fmt),
        ("Risky upside amount (tangency +1 std dev)", res.upside_amount if res.used_stocks else None, money_fmt),
        ("Risky downside amount (tangency -1 std dev)", res.downside_amount if res.used_stocks else None, money_fmt),
        ("Implied risk-aversion (gamma, CRRA)", res.gamma if res.used_stocks else None, "0.0000"),
    ]
    for label, val, fmt in rows:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=val)
        if fmt:
            c.number_format = fmt
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Final Allocation").font = bold
    ws.cell(row=r, column=1).fill = section_fill
    r += 1
    ws.cell(row=r, column=1, value="Asset").font = bold
    ws.cell(row=r, column=2, value="Weight").font = bold
    ws.cell(row=r, column=3, value="Amount (LKR)").font = bold
    for c in (1, 2, 3):
        ws.cell(row=r, column=c).fill = header_fill
        ws.cell(row=r, column=c).border = border
    r += 1

    ws.cell(row=r, column=1, value="G-sec (Treasury bill)")
    ws.cell(row=r, column=2, value=res.gsec_weight).number_format = pct_fmt
    ws.cell(row=r, column=3, value=res.gsec_weight * investment_amount).number_format = money_fmt
    r += 1
    for t, w in res.equity_weights.items():
        ws.cell(row=r, column=1, value=t)
        ws.cell(row=r, column=2, value=w).number_format = pct_fmt
        ws.cell(row=r, column=3, value=w * investment_amount).number_format = money_fmt
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Portfolio Summary").font = bold
    ws.cell(row=r, column=1).fill = section_fill
    r += 1
    ws.cell(row=r, column=1, value="Total expected return (horizon)")
    ws.cell(row=r, column=2, value=res.portfolio_expected_return_pct / 100).number_format = pct_fmt
    r += 1
    ws.cell(row=r, column=1, value="Total volatility (horizon)")
    ws.cell(row=r, column=2, value=res.portfolio_volatility_pct / 100).number_format = pct_fmt
    r += 1
    ws.cell(row=r, column=1, value="Expected portfolio value at horizon end")
    ws.cell(row=r, column=2, value=res.portfolio_expected_value).number_format = money_fmt
    r += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    args = parse_args()

    if args.tenor and args.amount is not None:
        tenor = args.tenor.strip().title()
        amount = args.amount
        preference = [t.strip().upper() for t in args.preference.split(",")] if args.preference else None
        name = args.investor_name
    else:
        tenor, amount, preference, name = ask_interactive()

    res = run_allocation(
        portfolio_xlsx=args.portfolio,
        dataset_xlsx=args.dataset,
        tenor_month=tenor,
        investment_amount=amount,
        preference_list=preference,
    )
    print_report(res, amount)

    safe_name = "".join(c if c.isalnum() else "_" for c in name)
    fname = f"Allocation_Report_{safe_name}_{tenor}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(args.outdir, fname)
    build_excel_report(res, amount, name, tenor, out_path)
    print(f"\nExcel report saved: {out_path}")


if __name__ == "__main__":
    main()
